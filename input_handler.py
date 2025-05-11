import os
from openpyxl.styles import PatternFill
import json
import openpyxl
from datetime import datetime  
# Hàm hỗ trợ cơ bản
def convert_to_int(value):
    """Chuyển đổi chuỗi số có dấu phẩy thành số nguyên, trả về 0 nếu lỗi hoặc thiếu."""
    try:
        return int(str(value).replace(',', ''))
    except (ValueError, TypeError):
        return 0

def to_float(value):
    """Chuyển đổi giá trị thành số thực, thay dấu phẩy bằng dấu chấm, trả về 0.0 nếu lỗi hoặc thiếu."""
    try:
        if isinstance(value, str):
            value = value.replace(',', '.')
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def is_float(value):
    """Kiểm tra xem giá trị có thể chuyển thành số thực không."""
    try:
        float(str(value).replace(',', '.'))
        return True
    except (ValueError, TypeError):
        return False

def calculate_diff_for_side(rows, i, side):
    """Tính chênh lệch khối lượng chờ mua hoặc chờ bán so với timestamp trước đó."""
    if i == 0:  # Dòng đầu tiên không có dòng trước để so sánh
        return [0] * 3
    prev_row = rows[i - 1]
    current_row = rows[i]
    
    diffs = []
    for j in range(3):
        current_key = current_row.get(f"Giá chờ {side} {j+1}", "")
        prev_key = prev_row.get(f"Giá chờ {side} {j+1}", "")
        current_vol = convert_to_int(current_row.get(f"Chờ {side} {j+1}", 0))
        prev_vol = convert_to_int(prev_row.get(f"Chờ {side} {j+1}", 0))
        
        if current_key == prev_key or not (is_float(current_key) and is_float(prev_key)):
            diff = current_vol - prev_vol
        else:
            prev_keys = [to_float(prev_row.get(f"Giá chờ {side} {k+1}", "")) for k in range(3)]
            prev_vols = [convert_to_int(prev_row.get(f"Chờ {side} {k+1}", 0)) for k in range(3)]
            current_key_float = to_float(current_key)
            if current_key_float == 0.0 or not prev_keys:
                diff = current_vol
            else:
                closest_idx = min(range(len(prev_keys)), key=lambda k: abs(prev_keys[k] - current_key_float) if prev_keys[k] != 0.0 else float('inf'))
                if abs(current_key_float - prev_keys[closest_idx]) <= 0.05:
                    prev_vol = prev_vols[closest_idx]
                    diff = current_vol - prev_vol
                else:
                    diff = current_vol
        diffs.append(diff)
    return diffs

def mark_price_change(all_rows):
    """Đánh dấu các hàng có thay đổi giá chờ mua hoặc chờ bán, chỉ khi dữ liệu đầy đủ."""
    for i in range(1, len(all_rows)):
        change = False
        direction = None
        has_missing_data = False
        
        for side in ['mua', 'bán']:
            for pos in range(1, 4):
                curr_val = all_rows[i].get(f"Giá chờ {side} {pos}", "")
                if not curr_val:
                    has_missing_data = True
                    break
            if has_missing_data:
                break
        
        if not has_missing_data:
            for side in ['mua', 'bán']:
                for pos in range(1, 4):
                    prev_val = all_rows[i-1].get(f"Giá chờ {side} {pos}", "")
                    curr_val = all_rows[i].get(f"Giá chờ {side} {pos}", "")
                    if prev_val != curr_val and prev_val and curr_val:
                        change = True
                        try:
                            if to_float(curr_val) < to_float(prev_val):
                                direction = "decrease"
                            else:
                                direction = "increase"
                        except:
                            direction = "unknown"
                        break
                if change:
                    break
        
        if change:
            all_rows[i]["Thay đổi bước giá"] = 1
            all_rows[i]["price_direction"] = direction
    return all_rows


def json_to_py(json_file: str): 
    """Xử lý file JSON và xuất ra Excel, điều chỉnh trật tự và bỏ qua dữ liệu lặp lại."""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Lỗi khi đọc file JSON {json_file}: {e}")
        return
    
    timestamps = sorted(data.keys())
    all_rows = []
    for timestamp in timestamps:
        entry = data[timestamp]
        zone = entry.get('zone', [{} for _ in range(3)])
        row = {
            'Thời gian': timestamp,
            'Giá': entry.get('Giá', {}).get('Lần 1', '') or '0',
            'Khối lượng': entry.get('KL', {}).get('Lần 1', '') or '0',
            'Loại': entry.get('M/B', {}).get('Lần 1', '') or '',
            'Khớp lần 1': entry.get('Khớp', {}).get('Lần 1', '') or ''
        }
        for i in range(3):
            zone_data = zone[i] if i < len(zone) else {}
            row[f'Chờ mua {i+1}'] = convert_to_int(zone_data.get(f'KL_mua {i+1}', '0'))
            row[f'Giá chờ mua {i+1}'] = zone_data.get(f'Gia_mua {i+1}', '') or ''
            row[f'Chờ bán {i+1}'] = convert_to_int(zone_data.get(f'KL_ban {i+1}', '0'))
            row[f'Giá chờ bán {i+1}'] = zone_data.get(f'Gia_ban {i+1}', '') or ''
        all_rows.append(row)
    
    # Tính chênh lệch
    for i in range(len(all_rows)):
        for side in ['mua', 'bán']:
            diffs = calculate_diff_for_side(all_rows, i, side)
            for j in range(3):
                all_rows[i][f"Tăng/giảm chờ {side} {j+1}"] = diffs[j]
    
    # Xử lý lặp lại trước khi sắp xếp ngược
    last_change = None
    for i in range(len(all_rows)):
        curr_row = all_rows[i]
        if i == 0:
            last_change = {
                'Giá': curr_row['Giá'],
                'Khối lượng': curr_row['Khối lượng'],
                'Loại': curr_row['Loại']
            }
        else:
            if (curr_row['Giá'] == last_change['Giá'] and
                curr_row['Khối lượng'] == last_change['Khối lượng'] and
                curr_row['Loại'] == last_change['Loại']):
                curr_row['Giá'] = ''
                curr_row['Khối lượng'] = ''
                curr_row['Loại'] = ''
            else:
                last_change = {
                    'Giá': curr_row['Giá'],
                    'Khối lượng': curr_row['Khối lượng'],
                    'Loại': curr_row['Loại']
                }
    
    # Sắp xếp ngược từ mới nhất đến cũ nhất
    all_rows = sorted(all_rows, key=lambda x: datetime.strptime(x['Thời gian'], "%H:%M:%S"), reverse=True)
    
    all_rows = mark_price_change(all_rows) 

    return all_rows 


def py_to_excel(all_rows: list, file_name:str): 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dữ liệu"
    
    headers = [
        "Thời gian", "Giá", "Khối lượng", "Loại",
        "Tăng/giảm chờ mua 1", "Tăng/giảm chờ bán 1",
        "Tăng/giảm chờ mua 2", "Tăng/giảm chờ bán 2",
        "Tăng/giảm chờ mua 3", "Tăng/giảm chờ bán 3",
        "Loại chờ mua mức 1 (C1)", "Loại chờ bán mức 1 (D1)",
        "Loại chờ mua mức 2 (C2)", "Loại chờ bán mức 2 (D2)",
        "Loại chờ mua mức 3 (C3)", "Loại chờ bán mức 3 (D3)",
        "Chờ mua 1", "Chờ bán 1", "Chờ mua 2", "Chờ bán 2",
        "Chờ mua 3", "Chờ bán 3",
        "Giá chờ mua 1", "Giá chờ bán 1",
        "Giá chờ mua 2", "Giá chờ bán 2",
        "Giá chờ mua 3", "Giá chờ bán 3",
        "Thay đổi bước giá"
    ]
    ws.append(headers)
    
    for row in all_rows:
        ws.append([row.get(header, '') for header in headers])
    
    color_decrease = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    color_increase = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for r_idx, row in enumerate(all_rows, start=2):
        if row.get("Thay đổi bước giá") == 1:
            direction = row.get("price_direction", "decrease")
            fill_color = color_decrease if direction == "decrease" else color_increase
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = fill_color 

        if row.get("Đánh dấu cụm") == 1:
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    stock_name = os.path.basename(file_name).split('.')[0]
    excel_file = os.path.join(os.path.dirname(file_name), f"{stock_name}.xlsx")
    wb.save(excel_file)
    print(f"Đã xuất ra Excel: {excel_file}")

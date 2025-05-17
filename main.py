import json
import os
from openpyxl.styles import PatternFill
import json
import openpyxl
from datetime import datetime    

def generate_next_color(initial_color: str = "000000") -> str:
    """
    Generates next color by adding an offset, ensuring good contrast with black text.
    Args:
        initial_color: Hex color string without '#' (default "000000")
    Returns:
        Next color as hex string with good readability
    """
    # Remove # if present and ensure 6 digits
    initial_color = initial_color.replace("#", "").zfill(6)
    
    # Convert hex to integer
    color_int = int(initial_color, 16)
    
    # Add offset (using golden ratio * 16^4 for good distribution)
    offset = int(0.618033988749895 * 65536)
    next_color = (color_int + offset) % 16777216
    
    # Extract RGB components
    r = (next_color >> 16) & 255
    g = (next_color >> 8) & 255
    b = next_color & 255
    
    # Ensure minimum brightness (adjust these values as needed)
    MIN_BRIGHTNESS = 90  # Increased minimum brightness
    r = max(r, MIN_BRIGHTNESS)
    g = max(g, MIN_BRIGHTNESS)
    b = max(b, MIN_BRIGHTNESS)
    
    # Combine back to hex
    return f"{r:02x}{g:02x}{b:02x}"

# Example usage:
# color = generate_next_color("FF0000")  # Starting from red
# pattern = PatternFill(start_color=color, end_color=color, fill_type="solid")

def to_mau_muc_gia(all_rows: str) -> dict:     
    price_field = ["Giá chờ mua 1", "Giá chờ mua 2", "Giá chờ mua 3", "Giá chờ bán 1", "Giá chờ bán 2", "Giá chờ bán 3"]
    used_price = set()
    map_price_to_color = {} 
    prev_color = "000000"
    for i in range(len(all_rows)):
        row = all_rows[i]    
        for field in price_field: 
            if field in row: 
                value = row[field]     

                if value == "ATC": 
                    continue 
                    
                if "Thay đổi bước giá" in all_rows[i]:
                    if all_rows[i]["Thay đổi bước giá"] == 1: 
                        # Nếu có thay đổi bước giá thì không tô màu
                        continue


                # new price 
                if value not in used_price: 
                    used_price.add(value)    
                    color = generate_next_color(prev_color) 
                    prev_color = color
                    map_price_to_color[value] = color     

                    # price color 
                    if "price color" not in all_rows[i]: 
                        all_rows[i]["price color"] = {}
                    all_rows[i]["price color"][field] = PatternFill(start_color=color , end_color=color , fill_type="solid") 

                    # weight color    
                    weight_field = field.replace("Giá ", "").capitalize()
                    all_rows[i]["price color"][weight_field] = PatternFill(start_color=color , end_color=color , fill_type="solid")

                # old price 
                else:  
                    color = map_price_to_color.get(value)
                    if "price color" not in all_rows[i]: 
                        all_rows[i]["price color"] = {}
                    all_rows[i]["price color"][field] = PatternFill(start_color=color , end_color=color , fill_type="solid")  

                    # weight color    
                    weight_field = field.replace("Giá ", "").capitalize()
                    all_rows[i]["price color"][weight_field] = PatternFill(start_color=color , end_color=color , fill_type="solid")

    return all_rows
                    

# Hàm hỗ trợ cơ bản
def convert_to_int(value):
    """Chuyển đổi chuỗi số có dấu phẩy thành số nguyên, trả về 0 nếu lỗi hoặc thiếu."""
    try:
        return int(str(value).replace(',', ''))
    except (ValueError, TypeError):
        return 0

def to_float(value):
    """Chuyển đổi giá trị thành số thực, thay dấu phẩy bằng dấu chấm, trả về 0.0 nếu lỗi hoặc thiếu."""
    try:
        if isinstance(value, str):
            value = value.replace(',', '.')
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def is_float(value):
    """Kiểm tra xem giá trị có thể chuyển thành số thực không."""
    try:
        float(str(value).replace(',', '.'))
        return True
    except (ValueError, TypeError):
        return False


def calculate_diff_for_side(rows, i, side):
    """Tính chênh lệch khối lượng chờ mua hoặc chờ bán so với timestamp trước đó."""
    if i == 0:  # Dòng đầu tiên không có dòng trước để so sánh
        return [0] * 3
    prev_row = rows[i - 1]
    current_row = rows[i]
    
    diffs = []
    for j in range(3):
        current_key = current_row.get(f"Giá chờ {side} {j+1}", "")
        prev_key = prev_row.get(f"Giá chờ {side} {j+1}", "")
        current_vol = convert_to_int(current_row.get(f"Chờ {side} {j+1}", 0))
        prev_vol = convert_to_int(prev_row.get(f"Chờ {side} {j+1}", 0))
        
        if current_key == prev_key or not (is_float(current_key) and is_float(prev_key)):
            diff = current_vol - prev_vol
        else:
            prev_keys = [to_float(prev_row.get(f"Giá chờ {side} {k+1}", "")) for k in range(3)]
            prev_vols = [convert_to_int(prev_row.get(f"Chờ {side} {k+1}", 0)) for k in range(3)]
            current_key_float = to_float(current_key)
            if current_key_float == 0.0 or not prev_keys:
                diff = current_vol
            else:
                closest_idx = min(range(len(prev_keys)), key=lambda k: abs(prev_keys[k] - current_key_float) if prev_keys[k] != 0.0 else float('inf'))
                if abs(current_key_float - prev_keys[closest_idx]) <= 0.05:
                    prev_vol = prev_vols[closest_idx]
                    diff = current_vol - prev_vol
                else:
                    diff = current_vol
        diffs.append(diff)
    return diffs 

def mark_price_change(all_rows):
    """Đánh dấu các hàng có thay đổi giá chờ mua hoặc chờ bán, chỉ khi dữ liệu đầy đủ."""
    for i in range(1, len(all_rows)):
        change = False
        direction = None
        has_missing_data = False
        
        for side in ['mua', 'bán']:
            for pos in range(1, 4):
                curr_val = all_rows[i].get(f"Giá chờ {side} {pos}", "")
                if not curr_val:
                    has_missing_data = True
                    break
            if has_missing_data:
                break
        
        if not has_missing_data:
            for side in ['mua', 'bán']:
                for pos in range(1, 4):
                    prev_val = all_rows[i-1].get(f"Giá chờ {side} {pos}", "")
                    curr_val = all_rows[i].get(f"Giá chờ {side} {pos}", "")
                    if prev_val != curr_val and prev_val and curr_val:
                        change = True
                        try:
                            if to_float(curr_val) < to_float(prev_val):
                                direction = "decrease"
                            else:
                                direction = "increase"
                        except:
                            direction = "unknown"
                        break
                if change:
                    break
        
        if change:
            all_rows[i]["Thay đổi bước giá"] = 1
            all_rows[i]["price_direction"] = direction
    return all_rows


def json_to_py(json_file: str): 
    """Xử lý file JSON và xuất ra Excel, điều chỉnh trật tự và bỏ qua dữ liệu lặp lại."""
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        print(f"Lỗi khi đọc file JSON {json_file}: {e}")
        return
    
    timestamps = sorted(data.keys())
    all_rows = []
    for timestamp in timestamps:
        entry = data[timestamp]
        zone = entry.get('zone', [{} for _ in range(3)])
        row = {
            'Thời gian': timestamp,
            'Giá': entry.get('Giá', {}).get('Lần 1', '') or '0',
            'Khối lượng': entry.get('KL', {}).get('Lần 1', '') or '0',
            'Loại': entry.get('M/B', {}).get('Lần 1', '') or '',
            'Khớp lần 1': entry.get('Khớp', {}).get('Lần 1', '') or ''
        }
        for i in range(3):
            zone_data = zone[i] if i < len(zone) else {}
            row[f'Chờ mua {i+1}'] = convert_to_int(zone_data.get(f'KL_mua {i+1}', '0'))
            row[f'Giá chờ mua {i+1}'] = zone_data.get(f'Gia_mua {i+1}', '') or ''
            row[f'Chờ bán {i+1}'] = convert_to_int(zone_data.get(f'KL_ban {i+1}', '0'))
            row[f'Giá chờ bán {i+1}'] = zone_data.get(f'Gia_ban {i+1}', '') or ''
        all_rows.append(row)
    
    # Tính chênh lệch
    for i in range(len(all_rows)):
        for side in ['mua', 'bán']:
            diffs = calculate_diff_for_side(all_rows, i, side)
            for j in range(3):
                all_rows[i][f"Tăng/giảm chờ {side} {j+1}"] = diffs[j]
    
    # Xử lý lặp lại trước khi sắp xếp ngược
    last_change = None
    for i in range(len(all_rows)):
        curr_row = all_rows[i]
        if i == 0:
            last_change = {
                'Giá': curr_row['Giá'],
                'Khối lượng': curr_row['Khối lượng'],
                'Loại': curr_row['Loại']
            }
        else:
            if (curr_row['Giá'] == last_change['Giá'] and
                curr_row['Khối lượng'] == last_change['Khối lượng'] and
                curr_row['Loại'] == last_change['Loại']):
                curr_row['Giá'] = ''
                curr_row['Khối lượng'] = ''
                curr_row['Loại'] = ''
            else:
                last_change = {
                    'Giá': curr_row['Giá'],
                    'Khối lượng': curr_row['Khối lượng'],
                    'Loại': curr_row['Loại']
                }
    
    # Sắp xếp ngược từ mới nhất đến cũ nhất
    all_rows = sorted(all_rows, key=lambda x: datetime.strptime(x['Thời gian'], "%H:%M:%S"), reverse=True)
    
    all_rows = mark_price_change(all_rows) 

    return all_rows 


def py_to_excel(all_rows: list, file_name:str): 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dữ liệu"
    
    #TODO: Tô màu cho các ô có giá trị giống nhau
    headers = [
        "Thời gian", "Giá", "Khối lượng", "Loại",
        "Tăng/giảm chờ mua 1", "Tăng/giảm chờ bán 1",
        "Tăng/giảm chờ mua 2", "Tăng/giảm chờ bán 2",
        "Tăng/giảm chờ mua 3", "Tăng/giảm chờ bán 3",
        "Loại chờ mua mức 1 (C1)", "Loại chờ bán mức 1 (D1)",
        "Loại chờ mua mức 2 (C2)", "Loại chờ bán mức 2 (D2)",
        "Loại chờ mua mức 3 (C3)", "Loại chờ bán mức 3 (D3)",
        "Chờ mua 1", "Chờ bán 1", "Chờ mua 2", "Chờ bán 2",
        "Chờ mua 3", "Chờ bán 3",
        "Giá chờ mua 1", "Giá chờ bán 1",
        "Giá chờ mua 2", "Giá chờ bán 2",
        "Giá chờ mua 3", "Giá chờ bán 3",
        "Thay đổi bước giá"
    ]
    ws.append(headers)
    
    for row in all_rows:
        ws.append([row.get(header, '') for header in headers])
    
    color_decrease = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    color_increase = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for r_idx, row in enumerate(all_rows, start=2):
        if row.get("Thay đổi bước giá") == 1:
            direction = row.get("price_direction", "decrease")
            fill_color = color_decrease if direction == "decrease" else color_increase
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = fill_color 

        # if row.get("Đánh dấu cụm") == 1:
        #     for c_idx in range(1, len(headers) + 1):
        #         ws.cell(row=r_idx, column=c_idx).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") 

        # Tô màu cho các ô  giá:     
        if row.get("price color") is not None:
            price_fields = ["Giá chờ mua 1", "Giá chờ mua 2", "Giá chờ mua 3", "Giá chờ bán 1", "Giá chờ bán 2", "Giá chờ bán 3"] 
            for field in price_fields:  
                if field in row["price color"]: 
                    color = row["price color"][field] 
                    if color is not None: 
                        ws.cell(row=r_idx, column=headers.index(field) + 1).fill = color

    # Tô  mauàu cho các ô khối lượng: 
        if row.get("price color") is not None:
            weight_fields = ["Chờ mua 1", "Chờ mua 2", "Chờ mua 3", "Chờ bán 1", "Chờ bán 2", "Chờ bán 3"] 
            for field in weight_fields:  
                if field in row["price color"]: 
                    color = row["price color"][field] 
                    if color is not None: 
                        ws.cell(row=r_idx, column=headers.index(field) + 1).fill = color
    
    stock_name = os.path.basename(file_name).split('.')[0]
    excel_file = os.path.join(os.path.dirname(file_name), f"{stock_name}.xlsx")
    wb.save(excel_file) 


    print(f"Đã xuất ra Excel: {excel_file}")

def is_giao_dich(row)->bool:   
    must_not_change_field = ["Khối lượng", "Giá", "Loại"]   

    for f in must_not_change_field:   
        if row[f] != "":
            return False  
    return True 

def is_lenh(row)->bool: 
    return not is_giao_dich(row) and row["Loại"] != ""

def is_in_time_interval(t1: dict, t2: dict, interval: int = 1) -> bool:  
    t1 = t1["Thời gian"]
    t2 = t2["Thời gian"]
    """
    Kiểm tra xem t1 và t2 có cách nhau interval giấy không 
    """
    time_format = "%H:%M:%S"
    t1_obj = datetime.strptime(t1, time_format)
    t2_obj = datetime.strptime(t2, time_format)
    
    # Tính khoảng cách giữa hai thời gian
    delta = abs((t2_obj - t1_obj).total_seconds())
    
    return delta <= interval

def get_list(l: list, idx:int):  
    if idx < 0:   
        return None  

    if idx >= len(l): 
        return None 

    return l[idx]

def lambda_in_time(t1:dict, interval:int = 1):  
    return lambda x: is_in_time_interval(t1, x, interval) or is_in_time_interval(x, t1, interval)

def get_row_at_idx(l: list, idx: int, c: list[callable]):  
    item = get_list(l,idx)  
    if item is None: 
        return None 
    else:
        for func in c:
            if not func(item): 
                return None
        return item

def gather_chunks(all_rows: list) -> list: 
    i = 0 
    chunk = []
    chunks:list[list] = [] 
    while i < len(all_rows):  
        row = all_rows[i] 
        if is_lenh(row): 
            next_gd = get_row_at_idx(all_rows, i + 1, [lambda_in_time(row), is_giao_dich]) 
            if next_gd:  
                chunk.append({"row": row, "idx": i})     

                chunk.append({"row": next_gd, "idx": i + 1}) 
                i += 1

        if len(chunk) > 0:
            chunks.append(chunk)
        chunk = []
        i += 1

    for c in chunks:  
        for row in c:  
            row["row"]["Đánh dấu cụm"] = True

    # chunks = [ 
    #     [{row: r1, idx: i1}, {row: r2, idx: i2}] 
    #     [{row:r3, idx: i3}, {row: r4, idx: i4}]
    # ]
         
    return chunks  


def edit_rows_based_on_chunks(all_rows: list, chunks: list) -> list:   
    new_list = all_rows.copy()  
    for chunk in chunks:  
        for row in chunk:    
            # for every row in chunk 
            # find row in original list 
            # edit it 
            for i in range(len(all_rows)):  
                if row.get("row") is None or row.get("row").get("Thời gian") is None: 
                    raise ValueError("Chunk format lỗi: ", row) 

                if all_rows[i]["Thời gian"] == row["row"]["Thời gian"]:   
                    # print("rewriting row: ", all_rows[i]["Thời gian"]) 
                    # print("row", row["row"])
                    new_list[i] = row["row"] 
                    break

    return new_list 


def fix_sync(all_rows: list) -> list:     
    new_rows = all_rows.copy()

    # Gom cụm 
    chunks = gather_chunks(all_rows)
    logs = [] 
    print("Tổng số cụm: ", len(chunks))  


    # Xử lý từng cụm
    new_chunks = []
    for i,_chunk in enumerate(chunks):     
        if len(_chunk) <= 1:  
            raise ValueError("Cụm không hợp lệ: ", _chunk)


        # Tính tổng    
        valid_perms = []
        valid_perms.append(_chunk)   
        valid_perms.append(_chunk[::-1]) 

        min_sum = float("inf") 
        best_perm_idx = -1   
        for i,p in enumerate(valid_perms):     
            sum = 0 

            # Lấy độ lệch từng hoán vị
            for j,row in enumerate(p):   
                row = row["row"]
                if j == len(p) - 1:
                    # Nếu là dòng cuối thì không cần tính
                    break

                delta = 0 

                # Dòng lệnh: Tăng/giảm = Chờ_mới - Chờ_cũ + KL_lệnh. 
                next_row = p[j+1]["row"]     
                if not is_giao_dich(row):    
                    delta = next_row["Chờ mua 1"] - row["Chờ mua 1"] + to_float(row["Khối lượng"])  
                else: 
                    # Dòng giao dịch: Tăng/giảm = Chờ_mới - Chờ_cũ (không cộng KL_lệnh).
                    delta = next_row["Chờ mua 1"] - row["Chờ mua 1"] 
                 
                sum += abs(delta) 

            # Chọn hoán vị tổng nhỏ nhất
            if sum < min_sum:  
                min_sum = sum
                best_perm_idx = i
        

        if best_perm_idx == -1:  
            raise ValueError("Không tìm thấy hoán vị hợp lệ") 
        
        best_perm = valid_perms[best_perm_idx]
        changed = best_perm_idx != 0
        sorted_time = sorted([x["row"]["Thời gian"] for x in best_perm], key=lambda x: datetime.strptime(x, "%H:%M:%S"), reverse=True)

        for i in range(len(best_perm)): 
            best_perm[i]["row"]["Thời gian"] = sorted_time[i] 

        # if not changed:  
        #     for i in range(len(best_perm)): 
        #         best_perm[i]["row"]["Đánh dấu cụm"] = False 
        new_chunks.append(best_perm) 


        logs.append({  
            "Timestamp ảnh hưởng ": [x["row"]["Thời gian"] for x in _chunk],
            "Có thay đổi: ": changed,  
            "Message: ": "Đã hoán vị cụm này" if changed else "Không có thay đổi",
            # "Thứ tự gốc: ": [x["Thời gian"] for x in _chunk],
            # "Thứ tự mới: ": [x["Thời gian"] for x in best_perm],
        })   

    new_rows = edit_rows_based_on_chunks(new_rows, new_chunks)
    
    return new_rows, logs

def main():
    # folder_path = input("Nhập đường dẫn thư mục chứa file JSON: ")
    # if not os.path.isdir(folder_path):
    #     print(f"Error: Folder {folder_path} does not exist")
    #     sys.exit(1)
    #
    # stock_code = input("Nhập mã chứng khoán (tên file JSON): ")
    # json_path = os.path.join(folder_path, f"{stock_code}.json")
    #
    # if not os.path.exists(json_path):
    #     print(f"Error: File {json_path} does not exist")
    #     sys.exit(1)
    #
    # if not json_path.endswith('.json'):
    #     print("Error: Input file must be a JSON file")
    #     sys.exit(1) 
    json_list = get_input_from_folder() 
    for json_path in json_list:
    
        print(f"Processing {json_path}...") 

        # đọc input từ file json thành list trong python 
        all_rows:list = json_to_py(json_path)           


        # Xử lý lỗi bất đồng bộ của data 
        fixed_rows, logs = fix_sync(all_rows)  
        fixed_rows = to_mau_muc_gia(fixed_rows)

        # in logs để debug
        # print("Log chỉnh trật tự: ",json.dumps(logs, indent=4, ensure_ascii=False))
        # Xuất ra excel 
        file_name = os.path.basename(json_path).split('.')[0] 
        py_to_excel(fixed_rows, file_name + "-fixed.xlsx")


        print("✅ Processing complete!") 

def test_gather_chunks():  
    all_rows:list = json_to_py("./TCH.json")           
    # Test gather_chunks function
    chunks = gather_chunks(all_rows) 

    new_list = edit_rows_based_on_chunks(all_rows, chunks)  
    py_to_excel(new_list, "./TCH_chunking.xlsx")

def get_input_from_folder():  
    # write a function to get a path to a folder  
    folder_path = input("Nhập đường dẫn thư mục chứa file JSON: ")  
    if not os.path.isdir(folder_path):
        print(f"Error: Folder {folder_path} does not exist")
        return None 
    
    # allow them to type in the name of the files  (no need to list files) 

    all_json_path = [] 
    try: 
        while True:
            filename = input("Enter JSON filename (or press CTRL-D to finish): ").strip()

            if filename == "":
                print("Error: Filename cannot be empty")
                continue
                
            if not filename.endswith('.json'):
                filename += '.json'
                
            if not os.path.exists(os.path.join(folder_path, filename)):
                print(f"Error: File '{filename}' does not exist in the folder '{folder_path}'")
                continue
            else: 
                print(f"File '{filename}' added to the list.")
                all_json_path.append(os.path.join(folder_path, filename))
    except EOFError:    
        print("Input ended. Processing files")
        pass  
    all_json_path = list(set(all_json_path))  # remove duplicates 
    return all_json_path


if __name__ == "__main__": 
    main()
